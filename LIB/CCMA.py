import pandas as pd
import pdfplumber
import re
import os
from tkinter.messagebox import showinfo
import openpyxl
from openpyxl.styles import PatternFill, Font

def CCMA():
    '''
    Función que lee todos los PDF de la carpeta 'PDF' y extrae los datos de los mismos. Luego hace una Tabla dinámica de control de los mismos.
    
    '''
    
    df = pd.DataFrame(columns=['Periodo', 'Concepto', 'Fecha de consulta', 'Monto', 'Archivo'])
    Archivos = []

    # Listar todos los PDF en 'PDF' con su ruta completa
    for root, dirs, files in os.walk('PDF'):
        for filename in files:
            #print(os.path.join(root, filename))
            Archivos.append(os.path.join(root, filename))

    # Filtrar los Archivos que no sean PDF
    Archivos = [x for x in Archivos if x.endswith(".pdf")]

    # Filtrar los Archivos que pesen menos de 10KB
    Archivos = [x for x in Archivos if os.path.getsize(x) > 10000]

    for i in Archivos:

        # Leer todas las páginas el PDF 'CCMA.pdf' de la carpeta 'PDF' e imprimir el resultado
        with pdfplumber.open(i) as pdf:
            pages = pdf.pages
            for page in pdf.pages:
                # Transformar en dataframe utilizando RegEx del grupo 1 y 4 de la siguiente expresión regular r'(\d+/\d+)\s(Saldo)\s(\d+/\d+/\d+)\s(\W?\d+.\d+\W?)'
                RegEx_Del_PDF = re.findall(r'(\d+/\d+)\s(Saldo)\s(\d+/\d+/\d+)\s(\W?\d+.\d+\W?)', page.extract_text() )

                # Transformar  RegEx_Del_PDF en dataframe donde la columna 0 es el 'Periodo', la columna 1 es el 'Concepto', la columna 2 es la 'Fecha de consulta' y la columna 3 es el monto y la columna 4 es el nombre del archivo PDF Leido
                for REg in RegEx_Del_PDF:
                    REg = list(REg)
                    REg.append(i)
                    df.loc[len(df)] = REg

    # Crear nuevas columnas a partir de la columna 'Archivo' y separar por '-' y eliminar los espacios finales e iniciales
    df['CUIT'] = df['Archivo'].str.split('-').str[3].str.strip()
    df['Cliente'] = df['Archivo'].str.split('-').str[4].str.strip().str.replace('.pdf', '')

    # Transormar la columna 'Monto' en float
    df['Monto'] = df['Monto'].str.replace(',', '').str.replace(',', '.').str.replace(')', '').str.replace('(', '-')
    df['Monto'] = df['Monto'].astype(float)

    # Crear una tabla dinámica con los datos de 'Cliente y 'Monto' y sumar los montos
    Td = pd.pivot_table(df, index=['Cliente'], values=['Monto'], aggfunc='sum')

    # Crear el Directorio 'Reporte' si no existe
    if not os.path.exists('Reporte'):
        os.makedirs('Reporte')

    # Exportar la tabla dinámica y el df a un archivo Excel
    with pd.ExcelWriter('Reporte/Reporte Consolidado CCMA.xlsx') as writer:
        Td.to_excel(writer, sheet_name='Reporte')
        df.to_excel(writer, sheet_name='Consolidado CCMA' , index=False)

    workbook = openpyxl.load_workbook('Reporte/Reporte Consolidado CCMA.xlsx')
    hoja1 = workbook['Reporte']  # Nombre de la hoja del DataFrame
    hoja2 = workbook['Consolidado CCMA']  # Nombre de la hoja del DataFrame

    # Darle formato a los Títulos de las columnas
    Fondotitulo = PatternFill(start_color='002060' , end_color='002060' ,  fill_type='solid')
    LetraColor = Font(color='FFFFFF')

    # Aplicar formato al encabezado
    for cell in hoja1[1]:
        cell.fill = Fondotitulo
        cell.font = LetraColor
    for cell in hoja2[1]:
        cell.fill = Fondotitulo
        cell.font = LetraColor

    # Autoajustar los anchos de las columnas según el contenido
    for column_cells in hoja1.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        hoja1.column_dimensions[column_cells[0].column_letter].width = length + 2
    for column_cells in hoja2.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        hoja2.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Agregar filtros de datos de ambas hojas
    hoja1.auto_filter.ref = hoja1.dimensions
    hoja2.auto_filter.ref = hoja2.dimensions

    # Guardar el archivo Excel
    workbook.save('Reporte/Reporte Consolidado CCMA.xlsx')

    # Mostrar mensaje de finalización
    showinfo("CCMA", "Se ha generado el archivo 'Reporte Consolidado CCMA.xlsx' en la carpeta de Reporte")


if __name__ == "__main__":
    CCMA()